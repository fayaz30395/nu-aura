from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json, os

wb = Workbook()
thin = Side(style='thin', color='000000')
border = Border(top=thin, left=thin, right=thin, bottom=thin)
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', fgColor='2B5797')
critical_fill = PatternFill('solid', fgColor='FF4444')
high_fill = PatternFill('solid', fgColor='FF8C00')
medium_fill = PatternFill('solid', fgColor='FFD700')
low_fill = PatternFill('solid', fgColor='90EE90')
fixed_fill = PatternFill('solid', fgColor='98FB98')
manual_fill = PatternFill('solid', fgColor='FFE4B5')
data_font = Font(name='Arial', size=10)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

def style_row(ws, row, cols, severity=None, status=None):
    fill = None
    if status == 'FIXED': fill = fixed_fill
    elif severity == 'critical': fill = critical_fill
    elif severity == 'high': fill = high_fill
    elif severity == 'medium': fill = medium_fill
    elif severity == 'low': fill = low_fill
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = data_font
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        if fill: cell.fill = fill

# ========== SUMMARY SHEET ==========
ws = wb.active
ws.title = 'Summary'
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 50

ws['A1'] = 'NU-AURA Full Platform Analysis Report'
ws['A1'].font = Font(name='Arial', bold=True, size=16, color='2B5797')
ws['A2'] = 'Date: 2026-03-27 | Convergence: 1 iteration'
ws['A2'].font = Font(name='Arial', size=11, color='666666')

r = 4
metrics = [
    ('Analysis Agents Deployed', '6', 'TypeScript, Backend, API, Integration, Security, UI/QA'),
    ('Total Findings', '52', 'Across all agents'),
    ('', '', ''),
    ('Critical Issues', '2', 'LazyInit in ApprovalEscalationJob, Cross-tenant AI usage query'),
    ('High Issues', '4', 'Missing audit, permission format, tenant filters'),
    ('Medium Issues', '12', 'Missing mutation error handlers, N+1, Kafka, scheduled jobs'),
    ('Low Issues', '30', 'Dev-gated console.logs (12), unused hooks (18)'),
    ('', '', ''),
    ('Auto-Fixed', '7', '6 mutation error handlers + 1 permission format normalization'),
    ('Manual Review Required', '13', 'LazyInit, tenant filters, missing tests, architecture'),
    ('Acceptable (No Action)', '32', 'Dev-gated logs, unused hooks (future use)'),
    ('', '', ''),
    ('TypeScript Compilation', 'PASS', 'Zero errors after fixes'),
    ('Backend Compilation', 'PASS', 'Zero errors'),
    ('Security Audit', '50/52 PASS', '2 warnings (permission format, coverage gap)'),
    ('SuperAdmin Bypass', '4/4 PASS', 'All locations verified'),
    ('OWASP Headers', '6/6 PASS', 'All headers present'),
    ('Password Policy', 'PASS', 'All rules enforced'),
]
for metric, value, detail in metrics:
    ws.cell(row=r, column=1, value=metric).font = Font(name='Arial', bold=True if metric else False, size=11)
    ws.cell(row=r, column=2, value=value).font = Font(name='Arial', size=11)
    ws.cell(row=r, column=3, value=detail).font = Font(name='Arial', size=10, color='666666')
    r += 1

# ========== AUTO-FIXED SHEET ==========
ws2 = wb.create_sheet('Auto-Fixed')
ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 45
ws2.column_dimensions['C'].width = 25
ws2.column_dimensions['D'].width = 50
ws2.column_dimensions['E'].width = 15

headers = ['Severity', 'File', 'Issue Type', 'Description', 'Status']
for c, h in enumerate(headers, 1):
    ws2.cell(row=1, column=c, value=h)
style_header(ws2, 1, 5)

fixed_items = [
    ('medium', 'frontend/app/attendance/shift-swap/page.tsx:100', 'missing-error-state', 'Added onError handler to submitMutation', 'FIXED'),
    ('medium', 'frontend/app/attendance/comp-off/page.tsx:72', 'missing-error-state', 'Added onError handler to requestMutation', 'FIXED'),
    ('medium', 'frontend/app/projects/resource-conflicts/page.tsx:42', 'missing-error-state', 'Added onError handler to scanMutation', 'FIXED'),
    ('medium', 'frontend/app/recruitment/job-boards/page.tsx:80', 'missing-error-state', 'Added onError handler to postMutation', 'FIXED'),
    ('medium', 'frontend/app/performance/pip/page.tsx:413', 'missing-error-state', 'Added onError handler to addCheckInMutation', 'FIXED'),
    ('medium', 'frontend/app/employees/change-requests/page.tsx:46', 'missing-error-state', 'Added onError handler to approveMutation', 'FIXED'),
    ('high', 'backend/.../controller/KekaImportController.java', 'permission-format', 'Normalized system.admin -> SYSTEM:ADMIN (7 annotations)', 'FIXED'),
]
for i, (sev, f, t, d, s) in enumerate(fixed_items, 2):
    ws2.cell(row=i, column=1, value=sev.upper())
    ws2.cell(row=i, column=2, value=f)
    ws2.cell(row=i, column=3, value=t)
    ws2.cell(row=i, column=4, value=d)
    ws2.cell(row=i, column=5, value=s)
    style_row(ws2, i, 5, status='FIXED')

# ========== MANUAL REVIEW SHEET ==========
ws3 = wb.create_sheet('Manual Review')
ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 50
ws3.column_dimensions['C'].width = 20
ws3.column_dimensions['D'].width = 60
ws3.column_dimensions['E'].width = 40

headers = ['Severity', 'File', 'Issue Type', 'Description', 'Recommended Action']
for c, h in enumerate(headers, 1):
    ws3.cell(row=1, column=c, value=h)
style_header(ws3, 1, 5)

manual_items = [
    ('critical', 'backend/.../scheduler/ApprovalEscalationJob.java:131', 'lazy-init', 'LazyInitializationException: accesses .getWorkflowExecution().getWorkflowDefinition() without @Transactional in @Scheduled method', 'Add @Transactional to processEscalationsForTenant() or use JOIN FETCH in query. Do NOT add @Transactional to @Scheduled method directly.'),
    ('critical', 'backend/.../repository/AiUsageLogRepository.java:17', 'missing-tenant-filter', 'Query sums tokens across ALL tenants without tenantId filter. Potential cross-tenant data leak.', 'Add AND a.tenantId = :tenantId to WHERE clause, or confirm this is intentional system-wide metric.'),
    ('high', 'backend/.../service/EmployeeService.java:575', 'missing-audit', 'deleteEmployee() terminates employees without explicit audit log entry.', 'Add auditService.log() call or verify EmployeeTerminatedEvent listener creates audit record.'),
    ('high', 'backend/src/main/java (35+ controllers)', 'missing-test', '35+ controllers lack corresponding test classes. JaCoCo 80% minimum.', 'Prioritize: Employee, Leave, Attendance, Payroll, Recruitment controllers.'),
    ('medium', 'backend/.../scheduler/AutoRegularizationScheduler.java:55', 'lazy-init', '@Scheduled autoRegularizeAttendance() lacks @Transactional. Helper has @Transactional but main method loops.', 'Verify transactional propagation. Consider wrapping loop body in separate @Transactional method.'),
    ('medium', 'backend/.../scheduler/AutoRegularizationScheduler.java:78', 'lazy-init', '@Scheduled autoApproveCompOff() accesses lazy relationships via compOffService.', 'Verify compOffService method is @Transactional. Add if missing.'),
    ('medium', 'backend/.../scheduler/ContractLifecycleScheduler.java:69', 'lazy-init', '@Scheduled processContractLifecycle() lacks @Transactional on main method.', 'Add @Transactional(readOnly=false) to main method or verify helpers handle session.'),
    ('medium', 'backend/.../scheduler/WorkflowEscalationScheduler.java:316', 'lazy-init', 'sendEscalationNotification() accesses lazy WorkflowExecution outside transaction.', 'Verify step is fully loaded before notification. Add eager fetch if needed.'),
    ('medium', 'backend/.../repository/AttendanceRecordRepository.java:46', 'missing-tenant-filter', 'Query filters by employeeId but not explicit tenantId. Relies on implicit FK isolation.', 'Add explicit tenantId for defense-in-depth.'),
    ('medium', 'backend/.../service/AttendanceImportService.java', 'n-plus-one', 'Loops over attendance records calling repo methods inside iteration.', 'Use batch operations (saveAll) instead of individual saves.'),
    ('medium', 'backend/.../kafka/producer/EventPublisher.java:299', 'kafka-unhandled', 'kafkaTemplate.send() uses .handle() but failed events only logged, not persisted to FailedKafkaEvent.', 'Verify DLT handler persists failures. If not, add FailedKafkaEvent persistence.'),
    ('medium', 'backend/.../notification/ScheduledNotificationService.java:53', 'lazy-init', '@Scheduled method lacks @Transactional. Low risk but inconsistent.', 'Add @Transactional for consistency.'),
    ('medium', 'ScheduledReportExecutionJob', 'config-issue', 'Runs every 1 minute. Potential performance bottleneck under load.', 'Consider increasing to 5-15 minute interval.'),
]
for i, (sev, f, t, d, a) in enumerate(manual_items, 2):
    ws3.cell(row=i, column=1, value=sev.upper())
    ws3.cell(row=i, column=2, value=f)
    ws3.cell(row=i, column=3, value=t)
    ws3.cell(row=i, column=4, value=d)
    ws3.cell(row=i, column=5, value=a)
    style_row(ws3, i, 5, severity=sev)

# ========== BY AGENT SHEET ==========
ws4 = wb.create_sheet('By Agent')
ws4.column_dimensions['A'].width = 25
ws4.column_dimensions['B'].width = 12
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 12
ws4.column_dimensions['E'].width = 12
ws4.column_dimensions['F'].width = 15

headers = ['Agent', 'Critical', 'High', 'Medium', 'Low', 'Status']
for c, h in enumerate(headers, 1):
    ws4.cell(row=1, column=c, value=h)
style_header(ws4, 1, 6)

agents = [
    ('TypeScript Health', 0, 0, 8, 30, 'COMPLETE'),
    ('Backend Code Health', 1, 2, 8, 1, 'COMPLETE'),
    ('API Contract Audit', 0, 1, 1, 0, 'COMPLETE (static)'),
    ('Integration Health', 0, 0, 1, 0, 'COMPLETE (static)'),
    ('Security & RBAC', 0, 1, 1, 0, 'COMPLETE'),
    ('UI/QA Browser Sweep', 0, 0, 0, 0, 'PARTIAL (needs dedicated run)'),
]
for i, (name, c, h, m, l, s) in enumerate(agents, 2):
    ws4.cell(row=i, column=1, value=name)
    ws4.cell(row=i, column=2, value=c)
    ws4.cell(row=i, column=3, value=h)
    ws4.cell(row=i, column=4, value=m)
    ws4.cell(row=i, column=5, value=l)
    ws4.cell(row=i, column=6, value=s)
    for col in range(1,7):
        ws4.cell(row=i, column=col).font = data_font
        ws4.cell(row=i, column=col).border = border

out = '/sessions/wizardly-eager-franklin/mnt/nu-aura/docs/superpowers/reports/nu-aura-analysis-report.xlsx'
wb.save(out)
print('Report saved to', out)
